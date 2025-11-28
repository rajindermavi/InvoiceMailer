from typing import List, Dict, Any, Iterable, Optional
from openpyxl import load_workbook

def iter_xlsx_rows_as_dicts(
    filepath: str,
    sheet_name: Optional[str] = None,
    header_row: int = 1,
):
    wb = load_workbook(filepath, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    header_cells = next(ws.iter_rows(
        min_row=header_row,
        max_row=header_row,
        values_only=True
    ))
    headers = [str(h) if h is not None else "" for h in header_cells]

    for row_cells in ws.iter_rows(
        min_row=header_row + 1,
        values_only=True
    ):
        if all(cell is None for cell in row_cells):
            continue
        yield {
            headers[i]: row_cells[i]
            for i in range(len(headers))
        }